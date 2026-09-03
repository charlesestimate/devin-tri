from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

F = "Arial"
HDR = PatternFill("solid", fgColor="1F3864"); HF = Font(name=F, bold=True, color="FFFFFF", size=10)
REQ = PatternFill("solid", fgColor="FFFF00"); OPT = PatternFill("solid", fgColor="FFF2CC")
EXF = Font(name=F, italic=True, color="7F7F7F", size=10); NF = Font(name=F, size=10)
thin = Side(style="thin", color="BFBFBF"); BR = Border(left=thin, right=thin, top=thin, bottom=thin)

# (column, required, enum or None, description, example)
S = {}
S["Persons"] = [
 ("full_name",1,None,"Legal full name as on payroll or contract","Mario Bautista"),
 ("aliases",0,None,"Other spellings seen in old records, separated by ; (prevents duplicates)","M. Bautista; Mario B."),
 ("population",1,"office|field|warehouse|consultant","Where this person works","field"),
 ("employment_basis",1,"employee|subcontractor_personnel|consultant","Basis of engagement","employee"),
 ("employer_party",0,None,"Legal name of employer if not Magnus (must exist in Parties)",""),
 ("signs_in",1,"yes|no","Will this person have a platform login","yes"),
 ("sign_in_email",0,None,"Email for login. Required when signs_in is yes","mario.bautista@magnus.example"),
 ("home_region",1,"Luzon|Bicol|Visayas|Mindanao","Home region","Luzon"),
 ("roles",0,None,"Role names separated by ; (must match Roles sheet). Assignment raises gate 24","Person In Charge"),
 ("capabilities",0,None,"Capability tags separated by ; e.g. working_at_height; electrical; lifting","working_at_height; electrical"),
 ("vouched_by",0,None,"Full name of Magnus person vouching. Required for subcontractor personnel",""),
 ("status",1,"active|suspended|departed","Current status","active"),
 ("source_reference",0,None,"Where this row came from (file, sheet, row)","HR masterlist 2026.xlsx row 12"),
]
S["Roles"] = [
 ("role_name",1,None,"Full role name, no abbreviations","Person In Charge"),
 ("department",1,None,"Department","Construction"),
 ("reports_to_role",0,None,"Role this reports to (empty only for Chief Executive Officer)","Project Manager"),
 ("is_approver",1,"yes|no","Does this role decide gates","no"),
 ("money_visibility",1,"none|cost|price|margin","Highest money level this role may see","none"),
 ("record_scope",1,"own|department|region|all","Which records this role may see","own"),
]
S["Parties"] = [
 ("legal_name",1,None,"Name on the contract, not the trading name","SolarTech Distribution Philippines Inc."),
 ("party_types",1,None,"One or more of client; asset_owner; offtaker; subcontractor; supplier; consultant; operations_and_maintenance_provider; other, separated by ;","supplier"),
 ("taxpayer_identification_number",0,None,"Required where invoiced or paid","000-000-000-000"),
 ("address",1,None,"Registered address","Biñan, Laguna"),
 ("is_related_party",1,"yes|no","Related to Magnus ownership","no"),
 ("accreditation_state",0,"accredited|provisional|suspended","Suppliers and subcontractors only","accredited"),
 ("insurance_expiry",0,None,"Date YYYY-MM-DD (subcontractors)",""),
 ("insurance_exclusions",0,None,"Exclusions written out, not attached",""),
 ("insurance_certificate_file",0,None,"File name of the certificate PDF you will send",""),
 ("bank_account_details",0,None,"Bank, account name, account number",""),
 ("categories_supplied",0,None,"For suppliers, separated by ;","panels; inverters"),
 ("currency",1,"PHP|USD|EUR|CNY","Billing currency","PHP"),
 ("source_reference",0,None,"Where this row came from","Supplier list.xlsx row 3"),
]
S["Accounts"] = [
 ("account_name",1,None,"Customer account (the organisation you sell to). Link to Parties by legal name in party_legal_name","Calamba Agro Industrial Corporation"),
 ("party_legal_name",1,None,"Legal name in Parties sheet","Calamba Agro Industrial Corporation"),
 ("industry",0,None,"Industry","Food processing"),
 ("active",1,"yes|no","Active account","yes"),
]
S["Sites"] = [
 ("site_name",1,None,"Short name used on screen","Calamba plant roof"),
 ("account_name",1,None,"Account this site belongs to (Accounts sheet)","Calamba Agro Industrial Corporation"),
 ("address",1,None,"Full address","Brgy. Real, Calamba, Laguna"),
 ("province",1,None,"Province. Region is derived from this, never typed","Laguna"),
 ("local_government_unit",1,None,"City or municipality","Calamba"),
 ("distribution_utility",1,None,"e.g. Meralco, SORECO II, NORECO II","Meralco"),
 ("host_party_legal_name",0,None,"Host party if different from account",""),
 ("emergency_nearest_hospital",0,None,"For the offline emergency card","Calamba Medical Center"),
 ("emergency_ambulance_number",0,None,"","0917 000 0000"),
 ("emergency_evacuation_point",0,None,"","Main gate parking"),
 ("emergency_client_contact",0,None,"Name and number","Engr. Cruz 0917 000 0001"),
]
S["Contacts"] = [
 ("full_name",1,None,"Contact person","Engr. Ana Cruz"),
 ("account_name",1,None,"Account (Accounts sheet)","Calamba Agro Industrial Corporation"),
 ("position",0,None,"","Plant Engineer"),
 ("electronic_mail_address",0,None,"","ana.cruz@example.com"),
 ("telephone_number",0,None,"","0917 000 0001"),
]
S["Items"] = [
 ("item_code",1,None,"Your internal code, unique","PNL-550-MONO"),
 ("description",1,None,"","550 watt monocrystalline panel"),
 ("specification",0,None,"","Bifacial, 144 half-cell"),
 ("unit_of_measure",1,None,"pc, m, kg, set, lot","pc"),
 ("is_serialised",1,"yes|no","Tracked by serial number","yes"),
 ("category",1,None,"panel; inverter; cable; mounting; electrical; safety; spare; other","panel"),
 ("reorder_point",0,None,"Operations and maintenance spares ONLY. Leave empty for project material",""),
]
S["Equipment"] = [
 ("description",1,None,"","Chain block 2 ton"),
 ("serial_number",1,None,"","CB-2024-017"),
 ("custodian_full_name",1,None,"A named person (Persons sheet), never a place","Jay Ramos"),
 ("certification_expiry",0,None,"Date YYYY-MM-DD for lifting gear, harnesses, testers","2027-03-31"),
 ("maintenance_due",0,None,"Date YYYY-MM-DD","2026-12-01"),
 ("current_project",0,None,"Project code if deployed now",""),
]
S["Locations"] = [
 ("name",1,None,"Warehouse or site store name","Laguna Warehouse"),
 ("location_type",1,"warehouse|site_stock","","warehouse"),
 ("region",1,"Luzon|Bicol|Visayas|Mindanao","Drives route class on transfers","Luzon"),
 ("custodian_full_name",1,None,"Accountable for count variances (Persons sheet)","Jay Ramos"),
 ("project_code",0,None,"Required for site_stock, empty for warehouse",""),
]
S["Opening Stock"] = [
 ("location_name",1,None,"Locations sheet","Laguna Warehouse"),
 ("item_code",1,None,"Items sheet","PNL-550-MONO"),
 ("quantity_counted",1,None,"Physical count on count date","320"),
 ("count_date",1,None,"Date YYYY-MM-DD of the physical count","2026-09-10"),
 ("counted_by_full_name",1,None,"","Jay Ramos"),
 ("serial_numbers",0,None,"For serialised items, separated by ;",""),
 ("unit_cost",0,None,"Cost per unit for valuation, PHP","9500"),
 ("condition",1,"good|quarantined","Damaged or disputed stock is loaded as quarantined","good"),
]
S["Projects"] = [
 ("project_code",1,None,"Your existing project code, unique","MRT-2025-014"),
 ("project_name",1,None,"","Calamba Agro 500 kWp rooftop"),
 ("client_legal_name",1,None,"Parties sheet","Calamba Agro Industrial Corporation"),
 ("site_name",1,None,"Sites sheet","Calamba plant roof"),
 ("capacity_kilowatt_peak",1,None,"Number","500"),
 ("mount_type",1,"rooftop|ground_mount|carport","Ground mount makes B11 gate B1","rooftop"),
 ("project_manager_full_name",1,None,"Persons sheet","Kidron Magnus"),
 ("director_full_name",1,None,"Approver on gates 6 and 8","Diego Fernandez"),
 ("person_in_charge_full_name",0,None,"Site lead","Mario Bautista"),
 ("status",1,"setup|active|suspended|turned_over","State as of cutover. Active requires the signed contract file","active"),
 ("mobilised",1,"yes|no","Already mobilised on site","yes"),
 ("insurance_certificate_file",0,None,"Required if mobilised and contract above 2,000,000","MRT-2025-014 CARI.pdf"),
 ("cshp_approved",1,"yes|no|not_required","Construction Safety and Health Program approval recorded","yes"),
 ("cshp_document_file",0,None,"","MRT-2025-014 CSHP approval.pdf"),
 ("permit_dependency",1,"parallel|prerequisite","Prerequisite blocks mobilisation until permit issued","parallel"),
 ("expected_permit_duration_days",0,None,"","45"),
 ("workday_counter_at_cutover",1,None,"Number of site workdays already elapsed. Without this every migrated project looks like it started on cutover day","62"),
 ("first_site_day",1,None,"Date YYYY-MM-DD of the first site day","2026-06-02"),
 ("planned_percentage_curve",0,None,"Date=percent pairs separated by ; from the contract programme","2026-06-30=10; 2026-08-31=45; 2026-11-30=100"),
 ("turnover_date",0,None,"Only for turned_over projects",""),
 ("source_reference",0,None,"","Project tracker.xlsx row 4"),
]
S["Contracts"] = [
 ("project_code",1,None,"Projects sheet","MRT-2025-014"),
 ("signed_document_file",1,None,"File name of the SIGNED PDF you will send. Not a draft","MRT-2025-014 signed contract.pdf"),
 ("contract_value",1,None,"Number, PHP unless currency says otherwise","28800000"),
 ("currency",1,"PHP|USD|EUR","","PHP"),
 ("date_signed",1,None,"YYYY-MM-DD","2026-05-15"),
 ("client_signatory",1,None,"","Juan Dela Cruz, President"),
 ("payment_terms_days",1,None,"","30"),
 ("retention_percentage",1,None,"Number, e.g. 10","10"),
 ("retention_reference_date_basis",1,"turnover|final_acceptance|last_invoice","","turnover"),
 ("retention_release_months",1,None,"","12"),
 ("warranty_months",1,None,"Read from the document","24"),
 ("counsel_review_state",1,"reviewed|proceeded_without_review","Gate 10 outcome as it happened","reviewed"),
 ("related_party",1,"yes|no","","no"),
 ("risk_terms_read",0,None,"Clause families present, separated by ; (liquidated damages; force majeure; warranty; termination; payment; variation; insurance; dispute)","liquidated damages; warranty; payment"),
]
S["Project Parties"] = [
 ("project_code",1,None,"","MRT-2025-014"),
 ("party_legal_name",1,None,"Parties sheet","Calamba Agro Industrial Corporation"),
 ("role",1,"client|host|asset_owner|offtaker|financier|epc_contractor|operations_and_maintenance_provider|landlord","One row per party per role. A sale usually gives the client four rows","client"),
]
S["Project Blocks"] = [
 ("project_code",1,None,"","MRT-2025-014"),
 ("block_code",1,"B0|B1|B2|B3|B4|B5|B6|B7|B8|B9|B10|B11|General Requirements|Battery Energy Storage System","Fixed spine. One row per block per project","B1"),
 ("included",1,"yes|no","B8, B11 and Battery Energy Storage System are conditional","yes"),
 ("block_cost_materials",0,None,"From costed bill of materials, PHP. Drives value weight. Empty for General Requirements","14200000"),
 ("block_cost_labour",0,None,"PHP","1800000"),
 ("state_at_cutover",1,"not_started|blocked_material|in_progress|complete|signed_off","","in_progress"),
 ("percent_complete_at_cutover",0,None,"Loaded position. Used only to seed the derived figure; never typed afterwards","55"),
 ("signed_off_date",0,None,"For signed_off blocks. B0 sign-off is what releases B1",""),
]
S["Bill of Materials"] = [
 ("project_code",1,None,"","MRT-2025-014"),
 ("block_code",1,None,"Every line belongs to a block","B1"),
 ("item_code",0,None,"Items sheet, if catalogued","PNL-550-MONO"),
 ("item_description",1,None,"","550 watt monocrystalline panel"),
 ("specification",0,None,"",""),
 ("quantity",1,None,"","910"),
 ("unit_of_measure",1,None,"","pc"),
 ("is_serialised",1,"yes|no","","yes"),
 ("unit_cost",1,None,"PHP","9500"),
 ("quantity_received_to_date",0,None,"","910"),
 ("purchase_order_reference",0,None,"Your PO number if bought","PO-2026-031"),
]
S["Open Purchase Orders"] = [
 ("purchase_order_reference",1,None,"Your PO number, unique","PO-2026-044"),
 ("project_code",1,None,"","MRT-2025-014"),
 ("supplier_legal_name",1,None,"Parties sheet","Nordwind Energy GmbH"),
 ("currency",1,"PHP|USD|EUR|CNY","","EUR"),
 ("exchange_rate_applied",0,None,"Required if not PHP","62.10"),
 ("exchange_rate_date",0,None,"YYYY-MM-DD","2026-08-12"),
 ("total_value",1,None,"In the PO currency","20800"),
 ("state",1,"issued|partially_received","Only open orders are migrated","issued"),
 ("expected_arrival_date",1,None,"YYYY-MM-DD","2026-10-05"),
 ("lines",1,None,"item_code:quantity:unit_cost separated by ;","INV-50K:4:5200"),
 ("approved_by_full_name",1,None,"Who approved it at the time (gate 4 or 5 as it happened)","Melanie Cruz"),
 ("approved_on",1,None,"YYYY-MM-DD","2026-08-12"),
]
S["Project Permits"] = [
 ("project_code",1,None,"","MRT-2025-014"),
 ("permit_type_name",1,None,"e.g. Building permit; Electrical permit; Fire safety evaluation clearance; Barangay clearance; Environmental compliance certificate; Distribution impact study; Permission to operate; Net metering agreement","Building permit"),
 ("group",1,"local_government_unit|environmental_and_safety|utility_interconnection","","local_government_unit"),
 ("issuing_body",1,None,"The specific office","Calamba City Engineering Office"),
 ("responsible_person_full_name",1,None,"","Austin Reyes"),
 ("mode",1,"parallel|prerequisite","","parallel"),
 ("state",1,"not_filed|filed|approved|expired","","filed"),
 ("date_filed",0,None,"","2026-07-01"),
 ("expected_approval_date",0,None,"Required once filed","2026-09-15"),
 ("date_approved",0,None,"",""),
 ("expiry_date",0,None,"",""),
]
S["Billing Milestones"] = [
 ("project_code",1,None,"","MRT-2025-014"),
 ("sequence",1,None,"1, 2, 3","1"),
 ("description",1,None,"","Down payment"),
 ("basis",1,"percentage_of_completion|milestone_event|fixed_date","","milestone_event"),
 ("amount_or_percentage",1,None,"Amount in PHP, or percent if basis is percentage_of_completion","8640000"),
 ("state",1,"not_due|claimable|claimed|certified|invoiced|paid","As of cutover","paid"),
 ("claimed_amount",0,None,"","8640000"),
 ("certified_amount",0,None,"","8640000"),
 ("invoice_reference",0,None,"","INV-2026-0102"),
 ("paid_date",0,None,"","2026-06-05"),
]
S["Service Agreements"] = [
 ("agreement_reference",1,None,"Your reference, unique","OM-2024-003"),
 ("site_name",1,None,"Sites sheet. Agreement attaches to the site, never a project","Lipa cold storage roof"),
 ("account_name",1,None,"Counterparty (Accounts sheet)","Lipa Cold Storage and Logistics Inc."),
 ("project_code",0,None,"Only where Magnus built the asset",""),
 ("agreement_document_file",1,None,"Signed PDF. Without it the agreement loads as draft and produces no charge","OM-2024-003 signed.pdf"),
 ("commencement_date",1,None,"YYYY-MM-DD","2024-11-01"),
 ("term_months",1,None,"","24"),
 ("scope_of_service",1,"preventive|corrective|preventive_and_corrective|monitoring_only","","preventive_and_corrective"),
 ("charge_basis",1,"fixed_periodic|per_kilowatt_peak|per_visit|hybrid","","fixed_periodic"),
 ("charge_amount",1,None,"PHP","45000"),
 ("charge_period",1,"monthly|quarterly|annual","","quarterly"),
 ("escalation_percentage",0,None,"","5"),
 ("escalation_month",0,None,"1 to 12","11"),
 ("state",1,"draft|active|terminated","","active"),
]
S["Service Level Terms"] = [
 ("agreement_reference",1,None,"","OM-2024-003"),
 ("severity",1,"total_outage|partial_outage|degraded|cosmetic","One row per agreement per severity. Active is refused with none","total_outage"),
 ("response_hours",1,None,"Elapsed hours to first response","4"),
 ("restoration_hours",1,None,"Elapsed hours to restored","48"),
]
S["Serviced Assets"] = [
 ("asset_reference",1,None,"Your reference, unique","AST-LIPA-01"),
 ("site_name",1,None,"","Lipa cold storage roof"),
 ("agreement_reference",0,None,"Empty if monitored but not contracted","OM-2024-003"),
 ("project_code",0,None,"Where Magnus built it",""),
 ("capacity_kilowatt_peak",1,None,"","300"),
 ("commissioning_date",1,None,"YYYY-MM-DD","2024-10-15"),
 ("warranty_expiry_date",0,None,"Typed ONLY where Magnus did not build it; otherwise derived from turnover plus warranty months","2026-10-15"),
 ("generation_monitoring_source",0,None,"e.g. SolarEdge portal; Huawei FusionSolar","Huawei FusionSolar"),
 ("generation_monitoring_access",0,None,"How Magnus accesses it (no passwords here)","Shared installer account"),
 ("tariff_per_kilowatt_hour",1,None,"PHP, site specific. Values lost generation","9.80"),
 ("state",1,"monitored|under_service|service_lapsed|decommissioned","","under_service"),
]
S["Asset Equipment"] = [
 ("asset_reference",1,None,"","AST-LIPA-01"),
 ("equipment_class",1,"panel|inverter|battery|mounting|monitoring|other","","inverter"),
 ("manufacturer",1,None,"","Huawei"),
 ("model",1,None,"","SUN2000-100KTL"),
 ("serial_number",0,None,"","HW100-2024-0091"),
 ("quantity",1,None,"","3"),
 ("installed_date",0,None,"","2024-10-10"),
 ("item_code",0,None,"Items sheet if catalogued",""),
]
S["Maintenance Plans"] = [
 ("agreement_reference",1,None,"","OM-2024-003"),
 ("activity",1,None,"","Panel cleaning and visual inspection"),
 ("interval_months",1,None,"","3"),
 ("first_due_date",1,None,"","2025-02-01"),
 ("next_due_date",1,None,"As of cutover","2026-11-01"),
 ("estimated_hours",0,None,"","16"),
 ("required_capability",0,None,"","working_at_height"),
 ("is_active",1,"yes|no","","yes"),
]
S["Open Work Orders"] = [
 ("work_order_reference",1,None,"","WO-2026-018"),
 ("asset_reference",1,None,"","AST-LIPA-01"),
 ("agreement_reference",0,None,"","OM-2024-003"),
 ("origin",1,"preventive|fault|client_request|warranty|monitoring_alert","","fault"),
 ("severity",1,"total_outage|partial_outage|degraded|cosmetic","","partial_outage"),
 ("raised_on",1,None,"YYYY-MM-DD HH:MM if known","2026-08-28 09:30"),
 ("reported_by_full_name",1,None,"","Engr. Ana Cruz"),
 ("assigned_to_full_name",0,None,"Must be a person who signs in","Jeferson Tolentino"),
 ("state",1,"raised|assigned|in_progress|restored","Only open orders are migrated","assigned"),
 ("fault_description",1,None,"","Inverter 2 tripping at midday"),
 ("is_billable",1,"yes|no","","no"),
]
S["Open Warranty Claims"] = [
 ("claim_reference",1,None,"","WC-2026-002"),
 ("asset_reference",1,None,"","AST-LIPA-01"),
 ("work_order_reference",1,None,"The work that found it","WO-2026-018"),
 ("claimed_against",1,"supplier|subcontractor|magnus","","supplier"),
 ("purchase_order_reference",0,None,"Required where claimed_against is supplier","PO-2024-019"),
 ("raised_on",1,None,"","2026-08-30"),
 ("claim_value",1,None,"PHP","180000"),
 ("evidence_file",1,None,"","WC-2026-002 photos.pdf"),
 ("state",1,"raised|submitted|accepted|rejected","","submitted"),
]

wb = Workbook()
rd = wb.active; rd.title = "README"
rows = [
 ("Magnus Renewable Tech Corp — migration load templates", None),
 ("Purpose", "One sheet per object type. Fill the yellow columns (required) and the light-yellow columns (optional). Row 2 of every sheet is an EXAMPLE in grey italics: overwrite it or delete it, never load it."),
 ("Dates", "YYYY-MM-DD. Money in numbers only, no commas, no peso sign. Yes or no, lower case."),
 ("Names that link sheets", "Persons by full_name · Parties by legal_name · Accounts by account_name · Sites by site_name · Items by item_code · Projects by project_code · Agreements by agreement_reference · Assets by asset_reference. Spell them identically everywhere; validation will reject a link that does not match."),
 ("Lists", "Where a cell allows several values, separate them with a semicolon and a space."),
 ("Files", "Where a column asks for a file name, put the file in a folder with exactly that name. Signed contracts and agreement documents are required for a project or agreement to load as active."),
 ("Order of loading", "1 Roles · 2 Persons · 3 Parties · 4 Accounts · 5 Sites · 6 Contacts · 7 Items · 8 Equipment · 9 Locations · 10 Opening Stock · 11 Projects · 12 Contracts · 13 Project Parties · 14 Project Blocks · 15 Bill of Materials · 16 Open Purchase Orders · 17 Project Permits · 18 Billing Milestones · 19 Service Agreements · 20 Service Level Terms · 21 Serviced Assets · 22 Asset Equipment · 23 Maintenance Plans · 24 Open Work Orders · 25 Open Warranty Claims"),
 ("Not migrated", "Finished projects, their documents and chat history, historical accounting, closed purchase orders, closed work orders, past payroll. Live and open records only."),
 ("Payroll rates and statutory tables", "Not in this workbook. They are entered on screen under dual control by Human Resource and Finance."),
 ("Never load", "Passwords, monitoring portal credentials, salaries of named people. Rates go on screen; access details go to the platform's document store, not this file."),
]
for i,(a,b) in enumerate(rows, start=1):
    rd.cell(i,1,a).font = Font(name=F, bold=True, size=11 if i==1 else 10)
    if b: c = rd.cell(i,2,b); c.font = NF; c.alignment = Alignment(wrap_text=True, vertical="top")
rd.column_dimensions["A"].width = 30; rd.column_dimensions["B"].width = 110
r0 = len(rows)+2
rd.cell(r0,1,"Sheet").font = HF; rd.cell(r0,1).fill = HDR
rd.cell(r0,2,"Rows you filled (write the count here when done)").font = HF; rd.cell(r0,2).fill = HDR
rd.cell(r0,3,"Required columns").font = HF; rd.cell(r0,3).fill = HDR
for k,(name,cols) in enumerate(S.items(), start=1):
    ws = wb.create_sheet(name[:31])
    for j,(col,req,enum,desc,ex) in enumerate(cols, start=1):
        c = ws.cell(1,j,col); c.font = HF; c.fill = HDR; c.border = BR; c.alignment = Alignment(wrap_text=True, vertical="top")
        note = ("REQUIRED. " if req else "Optional. ") + desc + (("\nAllowed: " + enum.replace("|", ", ")) if enum else "")
        c.comment = Comment(note, "Migration template")
        e = ws.cell(2,j,ex); e.font = EXF; e.border = BR
        fill = REQ if req else OPT
        for r in range(3, 203):
            cc = ws.cell(r,j); cc.fill = fill; cc.font = NF; cc.border = BR
        if enum and len(enum) < 250:
            dv = DataValidation(type="list", formula1='"' + enum.replace("|", ",") + '"', allow_blank=not req, showErrorMessage=True, errorTitle="Not allowed", error="Choose one of: " + enum.replace("|", ", "))
            ws.add_data_validation(dv); dv.add(f"{get_column_letter(j)}3:{get_column_letter(j)}202")
        ws.column_dimensions[get_column_letter(j)].width = max(14, min(42, len(col)+4))
    ws.row_dimensions[1].height = 30; ws.freeze_panes = "A3"
    rd.cell(r0+k,1,name).font = NF
    rd.cell(r0+k,2,"").font = NF
    rd.cell(r0+k,3,sum(1 for c in cols if c[1])).font = NF
rd.cell(r0+len(S)+1,1,"Legend").font = Font(name=F, bold=True)
a = rd.cell(r0+len(S)+2,1,"Yellow"); a.fill = REQ; rd.cell(r0+len(S)+2,2,"Required column").font = NF
b = rd.cell(r0+len(S)+3,1,"Light yellow"); b.fill = OPT; rd.cell(r0+len(S)+3,2,"Optional column").font = NF
c = rd.cell(r0+len(S)+4,1,"Grey italic"); c.font = EXF; rd.cell(r0+len(S)+4,2,"Example row, row 2 of each sheet. Hover a header for the rule.").font = NF
wb.save("MAGNUS-MIGRATION-TEMPLATES.xlsx")
print(len(S), "sheets")
